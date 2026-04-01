from dotenv import load_dotenv
load_dotenv()

from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Query, Body
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from pathlib import Path
import os
import tempfile
import re
from typing import Optional, List, Dict, Any, Tuple

# Existing imports
from parsers import nmmc
from parsers import mcgm_application_parser

# Local DB (replaces Supabase)
import db as local_db

app = FastAPI(
    docs_url="/api/docs",
    openapi_url="/api/openapi.json",
    redoc_url="/api/redoc",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "https://10.28.30.56",
        "http://10.28.30.56",
        "http://localhost:3000",
        "http://127.0.0.1:3000",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def startup():
    local_db.init_tables()


# -------------------------------------------------------------------
# PATHS (keep these consistent)
# -------------------------------------------------------------------
# Use project root (two levels up from this file) so local Windows
# setup and server setup both resolve paths correctly.
BASE_DIR = Path(__file__).resolve().parent.parent

ROUTES_DIR = BASE_DIR / "ROUTES"
MASTER_FILES_DIR = BASE_DIR / "master_files"

# Downloads (what you already fixed)
MASTER_BUDGET_PATH = BASE_DIR / "Master_Budget_BACKUP.xlsx"
MASTER_DN_PATH = BASE_DIR / "ROUTES/Route 45/CE_DN_MUMU25R045.xlsx"
MASTER_PO_PATH = MASTER_FILES_DIR / "master_po.xlsx"

# Route-analysis "budget DB" location (preferred)
# If you have the real old server file, keep it here:
MASTER_BUDGET_DB = MASTER_FILES_DIR / "master_budget.xlsx"

# Route report Excel template (override with ROUTE_REPORT_TEMPLATE absolute path if needed)
ROUTE_REPORT_TEMPLATE_DEFAULT = BASE_DIR / "MUM_Route_23_analysis - 2026-01-08 v2 SP.xlsx"

# -------------------------------------------------------------------
# HELPERS
# -------------------------------------------------------------------
def _pick_budget_file() -> Path:
    """
    Prefer master_files/master_budget.xlsx (old server-style "DB"),
    else fall back to BASE_DIR/Master_Budget_BACKUP.xlsx.
    """
    if MASTER_BUDGET_DB.exists():
        return MASTER_BUDGET_DB
    if MASTER_BUDGET_PATH.exists():
        return MASTER_BUDGET_PATH
    return MASTER_BUDGET_DB  # default (will fail with clear error)


def _normalize(s: Any) -> str:
    if s is None:
        return ""
    return str(s).strip()


def _extract_digits(s: str) -> str:
    """
    Extract a digit run from a string (used to match numeric route ids).
    Example:
      "MUM_Route_131" -> "131"
      "Route 131" -> "131"
      "131" -> "131"
    """
    s = _normalize(s)
    m = re.search(r"(\d+)", s)
    return m.group(1) if m else ""


def _excel_cached_value_xml_text(val: Any) -> str:
    """Format a number for OOXML <v> (no thousands separator)."""
    if val is None:
        return "0"
    try:
        x = float(val)
    except Exception:
        return "0"
    if x != x:  # NaN
        return "0"
    if abs(x - round(x)) < 1e-9:
        return str(int(round(x)))
    return f"{x:.15g}".rstrip("0").rstrip(".")


def _inject_ooxml_cached_v_after_formula(xml: str, cell_ref: str, v_text: str) -> str:
    """
    Insert or replace <v> immediately after </f> inside <c r="cell_ref" ...>...</c>.
    String-based so the workbook keeps its original namespace formatting (no ns0: rewrite).
    """
    needle = f'<c r="{cell_ref}"'
    start = xml.find(needle)
    if start < 0:
        return xml
    cell_close = xml.find("</c>", start)
    if cell_close < 0:
        return xml
    chunk = xml[start:cell_close]
    f_end_rel = chunk.find("</f>")
    if f_end_rel < 0:
        return xml
    ins = start + f_end_rel + len("</f>")
    tail_to_cell_end = xml[ins:cell_close]
    stripped = tail_to_cell_end.lstrip()
    if stripped.startswith("<v"):
        m = re.match(r"\s*<v[^>]*>.*?</v>", tail_to_cell_end, re.DOTALL)
        if m:
            end = ins + m.end()
            return xml[:ins] + f"<v>{v_text}</v>" + xml[end:]
        return xml
    return xml[:ins] + f"<v>{v_text}</v>" + xml[ins:]


def _patch_xlsx_formula_cached_values(xlsx_path: str, sheet_name: str, ref_to_value: Dict[str, float]) -> None:
    """
    openpyxl writes <f> for formulas but never writes <v> (cached result), so Excel
    shows blank cells until a full recalculation. Inject <v> next to each <f> for the
    given refs so the grid shows numbers immediately while keeping formulas.
    """
    import io
    import math
    import zipfile
    import xml.etree.ElementTree as ET

    path = Path(xlsx_path)
    if not path.exists() or not ref_to_value:
        return

    ref_map: Dict[str, str] = {}
    for k, v in ref_to_value.items():
        key = str(k).upper().replace("$", "")
        if v is None or (isinstance(v, float) and math.isnan(v)):
            continue
        try:
            ref_map[key] = _excel_cached_value_xml_text(v)
        except Exception:
            continue
    if not ref_map:
        return

    NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

    def _sheet_part_path(zin: zipfile.ZipFile) -> Optional[str]:
        try:
            root = ET.fromstring(zin.read("xl/workbook.xml"))
        except Exception:
            return None
        rid = None
        for el in root.iter():
            if el.tag == f"{{{NS_MAIN}}}sheet" and el.get("name") == sheet_name:
                rid = el.get(f"{{{NS_REL}}}id")
                break
        if not rid:
            return None
        try:
            rels = ET.fromstring(zin.read("xl/_rels/workbook.xml.rels"))
        except Exception:
            return None
        for el in rels:
            if el.tag.endswith("Relationship") and el.get("Id") == rid:
                tgt = (el.get("Target") or "").replace("\\", "/")
                if tgt.startswith("worksheets/"):
                    return "xl/" + tgt
                if "/worksheets/" in tgt:
                    return "xl" + tgt[tgt.index("/worksheets/") :].replace("/xl/", "xl/")
                return "xl/worksheets/" + tgt.rsplit("/", 1)[-1]
        return None

    with zipfile.ZipFile(path, "r") as zin:
        sheet_part = _sheet_part_path(zin)
        if not sheet_part or sheet_part not in zin.namelist():
            return
        try:
            xml = zin.read(sheet_part).decode("utf-8")
        except Exception:
            return

        for ref, vtxt in ref_map.items():
            xml = _inject_ooxml_cached_v_after_formula(xml, ref, vtxt)

        new_sheet_bytes = xml.encode("utf-8")

        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zout:
            for name in zin.namelist():
                if name == sheet_part:
                    zout.writestr(name, new_sheet_bytes)
                else:
                    zout.writestr(name, zin.read(name))

    path.write_bytes(buf.getvalue())


def _read_budget_df() -> "Any":
    """
    Reads the Excel budget file using pandas (openpyxl engine).
    Your backend venv already has pandas/openpyxl per requirements.txt.
    """
    try:
        import pandas as pd
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"pandas not available in backend venv. Install backend requirements. Error: {e}",
        )

    budget_file = _pick_budget_file()
    if not budget_file.exists():
        raise HTTPException(status_code=404, detail=f"Missing budget file: {budget_file}")

    # Try common sheet names; if none work, read first sheet.
    sheet_candidates = ["MasterBudget", "masterbudget", "Sheet1", "Budget", "MASTERBUDGET"]
    last_err = None
    for sh in sheet_candidates:
        try:
            df = pd.read_excel(str(budget_file), sheet_name=sh)
            return df
        except Exception as e:
            last_err = e

    # fallback: first sheet
    try:
        df = pd.read_excel(str(budget_file), sheet_name=0)
        return df
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Unable to read budget excel: {budget_file}. Last error: {last_err}. Final error: {e}",
        )


def _folder_route_ids_fallback() -> List[str]:
    """
    Fallback if budget file doesn't exist or doesn't have route_id_site_id.
    Extract numeric IDs from folder names.
    """
    route_ids = set()
    if ROUTES_DIR.exists():
        for p in ROUTES_DIR.glob("Route *"):
            m = re.search(r"Route\s+(\d+)", p.name, re.IGNORECASE)
            if m:
                route_ids.add(m.group(1))

        for p in ROUTES_DIR.glob("Coverage*"):
            m = re.search(r"Route\s*([0-9]+)", p.name, re.IGNORECASE)
            if m:
                route_ids.add(m.group(1))

    return sorted(route_ids, key=lambda x: int(x))


# -------------------------------------------------------------------
# ROUTE IDS (IMPORTANT: should match what route-analysis uses)
# -------------------------------------------------------------------
@app.get("/api/route-ids")
def get_route_ids():
    """
    Preferred: return unique route_id_site_id values from the budget Excel.
    Fallback: numeric folder-based IDs.
    """
    try:
        df = _read_budget_df()
        cols = [c.strip().lower() for c in df.columns.astype(str).tolist()]
        if "route_id_site_id" not in cols:
            # budget exists but doesn't have the column ? fallback to folders
            return {"route_ids": _folder_route_ids_fallback(), "source": "folders"}

        # get real column name as in df (preserve original)
        route_col = df.columns[cols.index("route_id_site_id")]
        series = df[route_col].dropna().astype(str).map(lambda x: x.strip())
        unique_vals = sorted(set([v for v in series.tolist() if v]), key=lambda x: (_extract_digits(x) or x, x))
        return {"route_ids": unique_vals, "source": str(_pick_budget_file())}
    except HTTPException:
        try:
            ids = local_db.get_budget_route_ids()
            if ids:
                return {"route_ids": sorted(ids), "source": "postgresql"}
        except Exception:
            pass
        return {"route_ids": _folder_route_ids_fallback(), "source": "folders"}
    except Exception:
        try:
            ids = local_db.get_budget_route_ids()
            if ids:
                return {"route_ids": sorted(ids), "source": "postgresql"}
        except Exception:
            pass
        return {"route_ids": _folder_route_ids_fallback(), "source": "folders"}


# -------------------------------------------------------------------
# ROUTE ANALYSIS (shared helper + endpoint)
# -------------------------------------------------------------------
def _route_analysis_rows(route_id_site_id: str) -> List[Dict[str, Any]]:
    """Returns budget rows for the given route_id_site_id (used by route-analysis and budget-by-route)."""
    df = _read_budget_df()
    cols_l = [c.strip().lower() for c in df.columns.astype(str).tolist()]
    if "route_id_site_id" not in cols_l:
        raise HTTPException(
            status_code=500,
            detail=f"Budget file {_pick_budget_file()} does not contain 'route_id_site_id' column.",
        )
    route_col = df.columns[cols_l.index("route_id_site_id")]
    q_raw = _normalize(route_id_site_id)
    q_low = q_raw.lower()
    q_digits = _extract_digits(q_raw)
    s = df[route_col].fillna("").astype(str).map(lambda x: x.strip())
    s_low = s.map(lambda x: x.lower())
    s_digits = s.map(_extract_digits)
    mask = (s_low == q_low)
    if q_digits and not mask.any():
        mask = (s_digits == q_digits)
    filtered = df[mask].copy()
    try:
        import pandas as pd
        filtered = filtered.where(pd.notnull(filtered), None)
        return filtered.to_dict(orient="records")
    except Exception:
        data = []
        for _, row in filtered.iterrows():
            rec = {str(k): (None if (v != v) else v) for k, v in row.items()}
            data.append(rec)
        return data


def _route_analysis_from_db(route_id_site_id: str) -> List[Dict[str, Any]]:
    """Fallback: get budget rows from PostgreSQL budget_master (exact then case-insensitive)."""
    try:
        sid = route_id_site_id.strip()
        rows = local_db.query_budget_by_site_id_all(sid)
        if not rows:
            rows = local_db.query_budget_by_route_id_insensitive(sid)
        return [dict(r) for r in rows] if rows else []
    except Exception:
        return []


@app.get("/api/route-analysis")
def route_analysis(route_id_site_id: str = Query(..., description="Route id (string or numeric)")):
    """
    Returns rows from budget filtered by route_id_site_id.
    Uses budget Excel if available; otherwise PostgreSQL budget_master.
    """
    data = []
    try:
        data = _route_analysis_rows(route_id_site_id)
    except HTTPException:
        data = _route_analysis_from_db(route_id_site_id)
    if not data:
        data = _route_analysis_from_db(route_id_site_id)
    return {"data": data, "rows": data}


# -------------------------------------------------------------------
# PO / BUDGET BY ROUTE (Developer Reference)
# -------------------------------------------------------------------
@app.get("/api/po-by-route")
def po_by_route(route_id_site_id: str = Query(..., description="Route id (string or numeric)")):
    """
    Returns a single PO row for the given route_id_site_id.
    Uses master_po.xlsx; first matching row or null.
    """
    try:
        import pandas as pd
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"pandas not available: {e}")
    if not MASTER_PO_PATH.exists():
        return {"data": [], "row": None}
    try:
        df = pd.read_excel(str(MASTER_PO_PATH), sheet_name=0)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Unable to read PO file: {e}")
    cols_l = [c.strip().lower() for c in df.columns.astype(str).tolist()]
    if "route_id_site_id" not in cols_l:
        return {"data": [], "row": None}
    route_col = df.columns[cols_l.index("route_id_site_id")]
    q_raw = _normalize(route_id_site_id)
    q_digits = _extract_digits(q_raw)
    s = df[route_col].fillna("").astype(str).map(lambda x: x.strip())
    s_low = s.map(lambda x: x.lower())
    s_digits = s.map(_extract_digits)
    mask = (s_low == q_raw.lower())
    if q_digits and not mask.any():
        mask = (s_digits == q_digits)
    filtered = df[mask]
    if filtered.empty:
        return {"data": [], "row": None}
    row = filtered.iloc[0].to_dict()
    row = {str(k): (None if (v != v) else v) for k, v in row.items()}
    return {"data": [row], "row": row}


@app.get("/api/budget-by-route")
def budget_by_route(route_id_site_id: str = Query(..., description="Route id (string or numeric)")):
    """
    Returns a single budget row (first from route-analysis) for the given route_id_site_id.
    """
    try:
        rows = _route_analysis_rows(route_id_site_id)
        row = rows[0] if rows else None
        return {"data": rows, "row": row}
    except HTTPException:
        return {"data": [], "row": None}
    except Exception:
        return {"data": [], "row": None}


# -------------------------------------------------------------------
# MASTER FILE DOWNLOADS (from PostgreSQL DB)
# -------------------------------------------------------------------
def _db_rows_to_excel_response(rows: List[Dict[str, Any]], filename: str):
    """Build Excel from DB rows and return as response."""
    import pandas as pd
    from io import BytesIO
    from fastapi.responses import Response
    if not rows:
        df = pd.DataFrame()
    else:
        # Drop internal id if present for cleaner export
        clean = [{k: v for k, v in r.items() if k != "id"} for r in rows]
        df = pd.DataFrame(clean)
    buf = BytesIO()
    df.to_excel(buf, index=False, sheet_name="Sheet1", engine="openpyxl")
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/download-master-po")
def download_master_po():
    rows = local_db.get_all_po_master()
    return _db_rows_to_excel_response(rows, "Master_PO.xlsx")


@app.get("/api/download-master-budget")
def download_master_budget():
    rows = local_db.get_all_budget_master()
    return _db_rows_to_excel_response(rows, "Master_Budget.xlsx")


@app.get("/api/download-master-dn")
def download_master_dn():
    rows = local_db.get_all_dn_master()
    return _db_rows_to_excel_response(rows, "Master_DN.xlsx")


# -------------------------------------------------------------------
# LOCAL DB API (replaces Supabase)
# -------------------------------------------------------------------
@app.get("/api/db/po-master/site-ids")
def api_po_site_ids():
    """Return distinct route_id_site_id from po_master."""
    ids = local_db.get_po_site_ids()
    return {"data": ids}


@app.get("/api/db/po-master")
def api_po_by_site(route_id_site_id: str = Query(..., alias="route_id_site_id")):
    row = local_db.query_po_by_site_id(route_id_site_id)
    return {"data": row, "error": None}


@app.get("/api/db/budget-master")
def api_budget_master(
    route_id_site_id: Optional[str] = Query(None, alias="route_id_site_id"),
    survey_ids: Optional[str] = Query(None),
    all_rows: Optional[bool] = Query(False, alias="all"),
):
    if route_id_site_id:
        if all_rows:
            rows = local_db.query_budget_by_site_id_all(route_id_site_id)
            return {"data": rows, "error": None}
        row = local_db.query_budget_by_site_id(route_id_site_id)
        return {"data": row, "error": None}
    if survey_ids:
        ids = [s.strip() for s in survey_ids.split(",") if s.strip()]
        rows = local_db.query_budget_by_survey_ids(ids)
        return {"data": rows, "error": None}
    raise HTTPException(status_code=400, detail="Provide route_id_site_id or survey_ids")


@app.post("/api/db/budget-master")
def api_budget_master_bulk(body: Dict[str, Any] = Body(...)):
    """Bulk replace: delete rows whose route_id_site_id not in list, then insert rows."""
    rows = body.get("rows", [])
    if not rows:
        return {"success": True, "message": "No rows to insert"}
    site_ids = [r.get("route_id_site_id") for r in rows if r.get("route_id_site_id")]
    local_db.budget_delete_not_in_site_ids(site_ids)
    local_db.budget_insert_many(rows)
    return {"success": True, "message": "Budget rows updated."}


@app.get("/api/db/dn-master/site-ids")
def api_dn_site_ids():
    """Return distinct route_id_site_id from dn_master."""
    ids = local_db.get_dn_site_ids()
    return {"data": ids}


@app.get("/api/db/dn-master")
def api_dn_master(
    dn_number: Optional[str] = Query(None),
    route_id_site_id: Optional[str] = Query(None, alias="route_id_site_id"),
):
    if dn_number:
        row = local_db.get_dn_by_number(dn_number)
        return {"data": row, "error": None}
    if route_id_site_id:
        rows = local_db.get_dn_by_route_id_site_id(route_id_site_id)
        return {"data": rows, "error": None}
    raise HTTPException(status_code=400, detail="Provide dn_number or route_id_site_id")


@app.post("/api/db/dn-master")
def api_dn_master_upsert(body: Dict[str, Any] = Body(...)):
    """Upsert a single dn_master row (e.g. permit fields)."""
    local_db.upsert_dn_master(body)
    return {"success": True}


# -------------------------------------------------------------------
# CLIENT PARSER V2 API
# -------------------------------------------------------------------
@app.get("/api/client-parser-v2/authorities")
def api_client_parser_authorities():
    """Return list of supported authority names for Client Parser V2."""
    from parsers.clientparserv2 import AUTHORITY_CONFIGS
    return {"authorities": list(AUTHORITY_CONFIGS.keys())}


@app.get("/api/client-parser-v2/dn-numbers")
def api_client_parser_dn_numbers():
    """Return distinct DN numbers from dn_master for Client Parser V2."""
    engine = local_db.get_engine()
    rows = local_db._run_sql(
        engine,
        "SELECT DISTINCT dn_number FROM dn_master "
        "WHERE dn_number IS NOT NULL AND dn_number != '' "
        "ORDER BY dn_number",
    )
    dn_numbers = [r.get("dn_number") for r in rows if r.get("dn_number")]
    return {"dn_numbers": dn_numbers}


@app.post("/api/client-parser-v2/validate-dn")
async def api_client_parser_validate_dn(dn_number: str = Form(...)):
    """Check if DN number exists in dn_master."""
    row = local_db.get_dn_by_number(dn_number.strip())
    return {"exists": row is not None}


@app.post("/api/client-parser-v2/unified")
async def api_client_parser_unified(
    dn_number: str = Form(...),
    authority: str = Form(...),
    output_type: str = Form("both"),
):
    """Run Client Parser V2 unified parser (non_refundable, sd, or both)."""
    from parsers.clientparserv2 import unified_parser
    result = unified_parser(dn_number.strip(), authority.strip().upper(), output_type)
    if "error" in result:
        raise HTTPException(status_code=500, detail=result["error"])
    return result


# -------------------------------------------------------------------
# ROUTE REPORT API (Excel + JSON)
# -------------------------------------------------------------------
def _build_route_report_rows(route_id_site_id: str) -> List[Dict[str, Any]]:
    """
    Build per-DN rows for a route, combining budget_master, po_master, dn_master.
    Keys are simple names for JSON; Excel uses table/column markers directly.
    """
    rid = (route_id_site_id or "").strip()
    if not rid:
        return []

    budget_rows = local_db.query_budget_by_site_id_all(rid)
    budget_row = budget_rows[0] if budget_rows else None
    po_row = local_db.query_po_by_site_id(rid)
    dn_rows = local_db.get_dn_by_route_id_site_id(rid)

    rows: List[Dict[str, Any]] = []
    for dn in dn_rows:
        row: Dict[str, Any] = {}
        # Core identifiers
        row["route_id_site_id"] = rid
        row["dn_number"] = dn.get("dn_number")
        row["dn_length_mtr"] = dn.get("dn_length_mtr")
        row["dn_received_date"] = dn.get("dn_received_date")
        row["actual_total_non_refundable"] = dn.get("actual_total_non_refundable")

        # Budget fields (example; extend as needed)
        if budget_row:
            row["budget_ce_length_mtr"] = budget_row.get("ce_length_mtr")
            row["budget_ri_cost_per_meter"] = budget_row.get("ri_cost_per_meter")
            row["budget_material_cost_per_meter"] = budget_row.get("material_cost_per_meter")
            row["budget_build_cost_per_meter"] = budget_row.get("build_cost_per_meter")
            row["budget_total_cost_without_deposit"] = budget_row.get("total_cost_without_deposit")

        # PO fields (example; extend as needed)
        if po_row:
            row["po_route_type"] = po_row.get("route_type")
            row["po_no_ip1"] = po_row.get("po_no_ip1")
            row["po_no_cobuild"] = po_row.get("po_no_cobuild")
            row["po_length_ip1"] = po_row.get("po_length_ip1")
            row["po_length_cobuild"] = po_row.get("po_length_cobuild")

        rows.append(row)

    return rows


def _build_route_report_workbook(
    route_id_site_id: str,
    modality: Optional[str] = None,
) -> Tuple["Any", Dict[str, float]]:
    """
    Open the template Excel and fill green cells whose values are
    table_name(column_name) by pulling from budget_master, po_master, dn_master.
    All formulas are left intact.

    Returns (workbook, formula_cached) — row 8 and per-DN F/G/J/K use Excel formulas
    built from actual DN summary row numbers (the template row 8 had #REF! placeholders
    that caused values to disappear after recalculation). formula_cached maps refs like
    A8 and F21 to numbers so /api/route-report/xlsx can inject OOXML <v> next to <f>
    (openpyxl does not write cached values for formulas).
    """
    from pathlib import Path as _Path
    import openpyxl as _openpyxl

    rid = (route_id_site_id or "").strip()
    if not rid:
        raise HTTPException(status_code=400, detail="route_id_site_id is required")

    # Fetch DB rows once (exact match, then case-insensitive route match)
    budget_rows = local_db.query_budget_by_site_id_all(rid)
    if not budget_rows:
        budget_rows = local_db.query_budget_by_route_id_insensitive(rid)
    budget_row = budget_rows[0] if budget_rows else None
    po_row = local_db.query_po_by_site_id(rid)
    if not po_row:
        po_row = local_db.query_po_by_route_id_insensitive(rid)
    planning_row = None
    try:
        planning_row = local_db.get_planning_tracker_by_route_id_site_id(rid)
    except Exception:
        planning_row = None
    dn_rows = local_db.get_dn_by_route_id_site_id(rid)
    if not dn_rows:
        dn_rows = local_db.get_dn_by_route_id_site_id_insensitive(rid)
    dn_rows = sorted(dn_rows, key=lambda x: str(x.get("dn_number") or ""))

    env_tpl = (os.environ.get("ROUTE_REPORT_TEMPLATE") or "").strip()
    template_path = _Path(env_tpl) if env_tpl else ROUTE_REPORT_TEMPLATE_DEFAULT
    if not template_path.exists():
        raise HTTPException(status_code=404, detail=f"Template not found: {template_path}")

    from openpyxl.utils import get_column_letter as _gcl

    wb = _openpyxl.load_workbook(template_path)
    sheet_name = "data-sheet"
    if sheet_name not in wb.sheetnames:
        raise HTTPException(status_code=500, detail=f"Sheet '{sheet_name}' not found in template")
    ws = wb[sheet_name]

    import re as _re

    def _coerce_numeric(value: Any) -> Any:
        """Convert Decimal / numeric-looking strings to float so openpyxl writes <v> not inlineStr."""
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return value
        if isinstance(value, str) and value.startswith("="):
            return value
        try:
            from decimal import Decimal
            if isinstance(value, Decimal):
                return float(value)
        except Exception:
            pass
        if isinstance(value, str):
            s = value.strip().replace(",", "")
            if not s:
                return value
            try:
                return float(s)
            except ValueError:
                return value
        try:
            return float(value)
        except (TypeError, ValueError):
            return value

    def _set_cell_safe(rr: int, cc: int, value: Any) -> None:
        """Write value; if cell is merged, write to merge anchor. Coerces to float when possible."""
        value = _coerce_numeric(value)
        cell = ws.cell(rr, cc)
        if cell.__class__.__name__ != "MergedCell":
            cell.value = value
            return
        coord = cell.coordinate
        for rng in ws.merged_cells.ranges:
            if coord in rng:
                ws.cell(rng.min_row, rng.min_col).value = value
                return

    # This template has explicit mapping rows:
    # - Row 2 contains green mapping cells for budget_master / po_master
    #   and row 4 is the corresponding data row.
    # - Row 18 contains green mapping cells for dn_master (dn_number, dn_length_mtr)
    #   and the DN "Approval Required" summary rows (21, 30, ...) are filled.
    map_pat = _re.compile(r"^[A-Za-z_]+\([A-Za-z0-9_]+\)$")

    sources_common = {
        "budget_master": budget_row or {},
        "po_master": po_row or {},
    }

    # --- 1) Fill budget/po mapping (row 2 -> row 4) ---
    mapping_row_budget = 2
    target_row_budget = 4
    for c in range(1, ws.max_column + 1):
        marker = ws.cell(mapping_row_budget, c).value
        if not (isinstance(marker, str) and map_pat.match(marker.strip())):
            continue
        table, col = marker.strip()[:-1].split("(", 1)
        table = table.strip()
        col = col.strip()
        src = sources_common.get(table, {})
        # Clear any template sample value first (but keep formulas if any)
        tgt_cell = ws.cell(target_row_budget, c)
        if not (isinstance(tgt_cell.value, str) and str(tgt_cell.value).startswith("=")):
            tgt_cell.value = None
        tgt_cell.value = _coerce_numeric(src.get(col))

    # --- 1b) Current template: row 2 has no table(column) markers — fill row 4 explicitly from DB ---
    def _sum_dn_survey_length(rows: List[Dict[str, Any]]) -> Optional[float]:
        s = 0.0
        any_v = False
        for dn in rows:
            v = dn.get("survey_length")
            if v is None or str(v).strip() == "":
                v = dn.get("application_length_mtr")
            if v is None or str(v).strip() == "":
                continue
            try:
                s += float(str(v).replace(",", ""))
                any_v = True
            except Exception:
                pass
        return s if any_v else None

    mod = (modality or "").strip()
    _set_cell_safe(4, 2, rid)
    if mod:
        _set_cell_safe(4, 3, mod)
    elif budget_row and budget_row.get("route_type"):
        _set_cell_safe(4, 3, budget_row.get("route_type"))
    elif po_row and po_row.get("route_type"):
        _set_cell_safe(4, 3, po_row.get("route_type"))

    if po_row:
        use_cobuild = mod and mod.lower() in ("co-build", "cobuild", "co build", "cobuilt", "co-built")
        if use_cobuild:
            _set_cell_safe(4, 4, po_row.get("po_length_cobuild"))
            _set_cell_safe(4, 15, po_row.get("po_no_cobuild"))
        else:
            _set_cell_safe(4, 4, po_row.get("po_length_ip1"))
            _set_cell_safe(4, 15, po_row.get("po_no_ip1"))

    e_val = None
    if budget_row and budget_row.get("ce_length_mtr") is not None:
        e_val = budget_row.get("ce_length_mtr")
    else:
        e_val = _sum_dn_survey_length(dn_rows)
    _set_cell_safe(4, 5, e_val)

    f_val = None
    if budget_row:
        f_val = budget_row.get("total_ri_amount")
        if f_val is None:
            f_val = budget_row.get("total_cost_without_deposit")
    _set_cell_safe(4, 6, f_val)

    if budget_row:
        # H4 = Budgeted Service/mtr, I4 = Budgeted Material/mtr (template row 3)
        _set_cell_safe(4, 8, budget_row.get("build_cost_per_meter"))
        _set_cell_safe(4, 9, budget_row.get("material_cost_per_meter"))
    # Same per-mtr values as H4/I4 for every DN summary row (H21/I21, H30/I30, …)
    h4_budget = budget_row.get("build_cost_per_meter") if budget_row else None
    i4_budget = budget_row.get("material_cost_per_meter") if budget_row else None

    if planning_row:
        _set_cell_safe(4, 13, planning_row.get("planning_date"))
    else:
        _set_cell_safe(4, 13, None)

    if budget_row and budget_row.get("category_type"):
        _set_cell_safe(4, 14, budget_row.get("category_type"))
    elif planning_row:
        _set_cell_safe(4, 14, planning_row.get("strategic_type"))
    else:
        _set_cell_safe(4, 14, None)

    # --- 2) DN block expansion + fill ---
    # Keep template structure exactly:
    # - row 20 is the label/header row
    # - row 21 is the data summary row for each DN block
    # - rows 22..28 are breakup/detail rows
    base_block_start = 20
    dn_block_height = 9  # rows 20..28 inclusive

    def _copy_block(src_start: int, src_height: int, dest_start: int) -> None:
        """
        Copy rows [src_start .. src_start+src_height-1] to start at dest_start,
        including styles and values. Formulas are shifted by row offset.
        """
        from copy import copy as _copy
        from openpyxl.formula.translate import Translator as _Translator

        row_offset = dest_start - src_start
        for r_src in range(src_start, src_start + src_height):
            r_dst = r_src + row_offset
            ws.row_dimensions[r_dst].height = ws.row_dimensions[r_src].height
            for c in range(1, ws.max_column + 1):
                src_cell = ws.cell(r_src, c)
                dst_cell = ws.cell(r_dst, c)
                # style
                dst_cell._style = _copy(src_cell._style)
                dst_cell.number_format = src_cell.number_format
                dst_cell.font = _copy(src_cell.font)
                dst_cell.fill = _copy(src_cell.fill)
                dst_cell.border = _copy(src_cell.border)
                dst_cell.alignment = _copy(src_cell.alignment)
                dst_cell.protection = _copy(src_cell.protection)
                dst_cell.comment = src_cell.comment

                v = src_cell.value
                if isinstance(v, str) and v.startswith("="):
                    # shift formulas to new row positions
                    try:
                        # Works across openpyxl versions: translate to destination coordinate.
                        dst_cell.value = _Translator(v, origin=src_cell.coordinate).translate_formula(dst_cell.coordinate)
                    except Exception:
                        # Fallback for versions supporting delta-style args.
                        try:
                            dst_cell.value = _Translator(v, origin=src_cell.coordinate).translate_formula(row_delta=row_offset, col_delta=0)
                        except Exception:
                            # Last resort: keep original formula text.
                            dst_cell.value = v
                else:
                    dst_cell.value = v

        # Copy merged ranges that are fully inside the source block.
        # Keeps row-20..28 structure identical across replicated DN blocks.
        for rng in list(ws.merged_cells.ranges):
            if rng.min_row < src_start or rng.max_row >= src_start + src_height:
                continue
            dst_min_row = rng.min_row + row_offset
            dst_max_row = rng.max_row + row_offset
            new_ref = f"{ws.cell(dst_min_row, rng.min_col).coordinate}:{ws.cell(dst_max_row, rng.max_col).coordinate}"
            if all(str(existing) != new_ref for existing in ws.merged_cells.ranges):
                ws.merge_cells(new_ref)

    if ws.cell(base_block_start, 1).value is None and ws.cell(base_block_start, 2).value is None:
        raise HTTPException(status_code=500, detail="Template row 20 DN block not found in data-sheet.")

    needed_blocks = max(1, len(dn_rows))  # keep one visible empty block when no DN rows exist
    existing_blocks = 1
    if needed_blocks > existing_blocks:
        insert_at = base_block_start + dn_block_height
        for _ in range(needed_blocks - existing_blocks):
            ws.insert_rows(insert_at, amount=dn_block_height)
            _copy_block(base_block_start, dn_block_height, insert_at)
            insert_at += dn_block_height

    dn_summary_rows = [base_block_start + 1 + i * dn_block_height for i in range(needed_blocks)]

    # DN mapping row defines which columns to fill in the summary rows
    mapping_row_dn = 18
    dn_col_mappings: Dict[int, str] = {}  # excel_col -> dn_master column
    for c in range(1, ws.max_column + 1):
        marker = ws.cell(mapping_row_dn, c).value
        if not (isinstance(marker, str) and map_pat.match(marker.strip())):
            continue
        table, col = marker.strip()[:-1].split("(", 1)
        table = table.strip()
        col = col.strip()
        if table == "dn_master":
            dn_col_mappings[c] = col

    # Summary columns filled by budget/RI math — do not let dn_master(row 18) mappings overwrite these.
    _DN_SUMMARY_PROTECT_COLS = {1, 2, 4, 5, 6, 7, 8, 9, 10, 11, 15}

    # Cached numeric results for OOXML <v> (openpyxl omits <v> for formulas).
    dn_formula_cached: Dict[str, float] = {}

    # Aggregates for row 8 cached values (<v>) — openpyxl does not write formula results.
    row8_sum_d = 0.0
    row8_sum_e = 0.0
    row8_sum_f = 0.0
    row8_sum_o = 0.0

    # Fill each DN into its own replicated 20..28 block.
    for idx, dn in enumerate(dn_rows):
        if idx >= len(dn_summary_rows):
            break
        r = dn_summary_rows[idx]
        # DN data cells in the summary row (as per template)
        # A{r} -> serial number (row 21 / 30 / 39...)
        # B{r} -> dn_master(dn_number), D{r} -> dn_master(dn_length_mtr)
        # Clear sample values first (preserve formulas elsewhere)
        _set_cell_safe(r, 1, None)
        _set_cell_safe(r, 2, None)
        _set_cell_safe(r, 4, None)
        _set_cell_safe(r, 1, idx + 1)
        _set_cell_safe(r, 2, dn.get("dn_number"))
        _set_cell_safe(r, 4, dn.get("dn_length_mtr"))
        _sv = dn.get("survey_length")
        if _sv is None or str(_sv).strip() == "":
            _sv = dn.get("application_length_mtr")
        _set_cell_safe(r, 5, _sv if _sv is not None and str(_sv).strip() != "" else None)

        def _split_by_comma(v: Any) -> List[str]:
            if v is None:
                return []
            s = str(v).strip()
            if not s:
                return []
            import re as _re_local
            return [x.strip() for x in _re_local.split(r"[,+]", s) if x.strip()]

        def _split_generic(v: Any) -> List[str]:
            if v is None:
                return []
            s = str(v).strip()
            if not s:
                return []
            import re as _re_local
            return [x.strip() for x in _re_local.split(r"[,/+]", s) if x.strip()]

        surfaces = _split_by_comma(dn.get("surface"))
        lengths = _split_generic(dn.get("surface_wise_length"))
        rates = _split_generic(dn.get("surface_wise_ri_amount"))
        factors = _split_generic(dn.get("surface_wise_multiplication_factor"))
        # Keep exact template structure in every DN block:
        # row r+2 = RI Charges line, r+3 = Ground Rent, r+4 = Admin, r+5 = Access/total
        ri_start = r + 2
        ri_end = ri_start
        ground_row = r + 3
        admin_row = r + 4
        access_row = r + 5

        def _to_float(x: Any) -> float:
            if x is None or x == "":
                return 0.0
            try:
                return float(str(x).replace(",", ""))
            except Exception:
                return 0.0

        try:
            _e_sur = float(str(_sv).replace(",", "")) if _sv is not None and str(_sv).strip() != "" else 0.0
        except Exception:
            _e_sur = 0.0
        d_len = _to_float(dn.get("dn_length_mtr"))

        # Fill dynamic per-surface lines in G/H/I/J and line total in K.
        for rr in range(ri_start, ri_end + 1):
            _set_cell_safe(rr, 7, None)  # G
            _set_cell_safe(rr, 8, None)  # H
            _set_cell_safe(rr, 9, None)  # I
            _set_cell_safe(rr, 10, None)  # J
            _set_cell_safe(rr, 11, None)  # K
            # Keep "RI Charges" label only on first row.
            if rr == ri_start:
                _set_cell_safe(rr, 5, "RI Charges")
            else:
                _set_cell_safe(rr, 5, None)
                _set_cell_safe(rr, 4, None)
                # Avoid repeated RI amount value in F (e.g. F24, F25...)
                _set_cell_safe(rr, 6, None)

        # Show first available surface details in the fixed RI row.
        if surfaces:
            _set_cell_safe(ri_start, 7, surfaces[0])
        if lengths:
            _set_cell_safe(ri_start, 8, lengths[0])
        if rates:
            _set_cell_safe(ri_start, 9, rates[0])
        if factors:
            f_raw = factors[0]
            try:
                f_num = float(str(f_raw).strip())
                _set_cell_safe(ri_start, 10, 1 if f_num == 0 else f_raw)
            except Exception:
                _set_cell_safe(ri_start, 10, f_raw)
        elif ws.cell(ri_start, 10).value in (None, ""):
            _set_cell_safe(ri_start, 10, 1)

        # Compute RI total from all surface-wise entries from dn_master while preserving row layout.
        k_row_totals: List[float] = []
        n_terms = max(len(lengths), len(rates), len(factors), 1)
        for i in range(n_terms):
            l = _to_float(lengths[i]) if i < len(lengths) else (0.0 if i > 0 else _to_float(ws.cell(ri_start, 8).value))
            rr_v = _to_float(rates[i]) if i < len(rates) else (0.0 if i > 0 else _to_float(ws.cell(ri_start, 9).value))
            ff = _to_float(factors[i]) if i < len(factors) else (1.0 if i == 0 else 0.0)
            if ff == 0:
                ff = 1.0
            k_row_totals.append(l * rr_v * ff)

        line_total_display = k_row_totals[0] if k_row_totals else 0.0
        _set_cell_safe(ri_start, 11, line_total_display if line_total_display != 0 else None)

        # Keep breakup lines below dynamic RI block.
        _set_cell_safe(ground_row, 6, dn.get("ground_rent"))
        _set_cell_safe(admin_row, 6, dn.get("administrative_charge"))
        if ws.cell(access_row, 6).value is None:
            _set_cell_safe(access_row, 6, dn.get("access_charges", 0) if isinstance(dn, dict) else 0)

        # Total K row should shift downward based on surface rows.
        # For default 3 surfaces, this remains K26; otherwise moves accordingly.
        k_total = sum(k_row_totals)
        _set_cell_safe(access_row, 11, k_total if k_total != 0 else None)
        # RI charges row in F points to the shifted K total.
        _set_cell_safe(ri_start, 6, k_total if k_total != 0 else None)

        k_numeric = k_total
        f_total = (
            k_numeric
            + _to_float(ws.cell(ground_row, 6).value)
            + _to_float(ws.cell(admin_row, 6).value)
            + _to_float(ws.cell(access_row, 6).value)
        )
        # F/G/J/K: same layout as template — formulas over breakup rows so Excel shows formulas, not #REF!.
        _set_cell_safe(r, 6, f"=SUM(F{r+2}:F{r+5})")
        h4_cell = ws.cell(4, 8).value
        i4_cell = ws.cell(4, 9).value
        h_val = h4_cell if h4_cell not in (None, "") else h4_budget
        i_val = i4_cell if i4_cell not in (None, "") else i4_budget
        g_dn = (f_total / d_len) if d_len else None
        h_dn = _to_float(h_val)
        i_dn = _to_float(i_val)
        j_dn = (g_dn + h_dn + i_dn) if d_len else None
        k_dn = (j_dn * d_len) if d_len else f_total
        _set_cell_safe(r, 7, f"=IF(D{r}=0,0,F{r}/D{r})")
        _set_cell_safe(r, 8, h_val)
        _set_cell_safe(r, 9, i_val)
        _set_cell_safe(r, 10, f"=G{r}+H{r}+I{r}")
        _set_cell_safe(r, 11, f"=J{r}*D{r}")
        dn_formula_cached[f"{_gcl(6)}{r}"] = float(f_total)
        dn_formula_cached[f"{_gcl(7)}{r}"] = float(g_dn) if g_dn is not None else 0.0
        dn_formula_cached[f"{_gcl(10)}{r}"] = float(j_dn) if j_dn is not None else 0.0
        dn_formula_cached[f"{_gcl(11)}{r}"] = float(k_dn) if k_dn is not None else 0.0
        dep_dn = _to_float(dn.get("deposit"))
        _set_cell_safe(r, 15, dep_dn if dep_dn else None)

        row8_sum_d += d_len
        row8_sum_e += _e_sur
        row8_sum_f += f_total
        row8_sum_o += dep_dn

        # Row 18 dn_master(...) markers: fill extra columns only (not G–K summary math).
        for c, dn_col in dn_col_mappings.items():
            if c in _DN_SUMMARY_PROTECT_COLS:
                continue
            _set_cell_safe(r, c, dn.get(dn_col))

    # --- 3) Row 8 totals — replace template #REF! sums with formulas over real DN summary rows only. ---
    rows_r = [dn_summary_rows[i] for i in range(len(dn_rows))] if dn_rows else [dn_summary_rows[0]]

    def _nf_r8(x: Any) -> float:
        if x is None or x == "":
            return 0.0
        try:
            return float(str(x).replace(",", ""))
        except Exception:
            return 0.0

    d8c = row8_sum_d
    e8c = row8_sum_e
    f8c = row8_sum_f
    o8c = row8_sum_o
    g8c = (f8c / d8c) if d8c else 0.0
    h8c = 0.0
    i8c = 0.0
    if d8c > 0 and rows_r:
        for rr in rows_r:
            h8c += _nf_r8(ws.cell(rr, 8).value) * _nf_r8(ws.cell(rr, 4).value)
            i8c += _nf_r8(ws.cell(rr, 9).value) * _nf_r8(ws.cell(rr, 4).value)
        h8c /= d8c
        i8c /= d8c
    j8c = g8c + h8c + i8c
    k8c = j8c * d8c
    a8c = float(len(dn_rows))

    ev = _nf_r8(e_val)
    fv = _nf_r8(f_val)
    h4b = _nf_r8(budget_row.get("build_cost_per_meter")) if budget_row else 0.0
    i4b = _nf_r8(budget_row.get("material_cost_per_meter")) if budget_row else 0.0
    g4v = (fv / ev) if ev else 0.0
    j4v = g4v + h4b + i4b
    l8c = j4v * d8c
    m8c = (j4v * d8c) - k8c
    n8c = (m8c / d8c) if d8c else 0.0
    p8c = f8c + o8c

    def _csv_col(col: int) -> str:
        return ",".join(f"{_gcl(col)}{r}" for r in rows_r)

    _set_cell_safe(8, 1, f"=COUNTA({_csv_col(2)})")
    _set_cell_safe(8, 4, f"=SUM({_csv_col(4)})")
    _set_cell_safe(8, 5, f"=SUM({_csv_col(5)})")
    _set_cell_safe(8, 6, f"=SUM({_csv_col(6)})")
    _set_cell_safe(8, 7, "=IF($D$8=0,0,$F$8/$D$8)")
    if rows_r:
        h_num = "+".join(f"H{r}*$D{r}" for r in rows_r)
        i_num = "+".join(f"I{r}*$D{r}" for r in rows_r)
        _set_cell_safe(8, 8, f"=IF($D$8=0,0,({h_num})/$D$8)")
        _set_cell_safe(8, 9, f"=IF($D$8=0,0,({i_num})/$D$8)")
    else:
        _set_cell_safe(8, 8, 0.0)
        _set_cell_safe(8, 9, 0.0)
    _set_cell_safe(8, 10, "=$G$8+$H$8+$I$8")
    _set_cell_safe(8, 11, "=$J$8*$D$8")
    _set_cell_safe(8, 12, "=IF($E$4=0,0,L$4/E$4*$D$8)")
    _set_cell_safe(8, 13, "=$J$4*$D$8-$K$8")
    _set_cell_safe(8, 14, "=IF($D$8=0,0,$M$8/$D$8)")
    _set_cell_safe(8, 15, f"=SUM({_csv_col(15)})")
    _set_cell_safe(8, 16, "=$F$8+$O$8")

    row8_cached: Dict[str, float] = {
        "A8": a8c,
        "D8": d8c,
        "E8": e8c,
        "F8": f8c,
        "G8": g8c,
        "H8": h8c,
        "I8": i8c,
        "J8": j8c,
        "K8": k8c,
        "L8": l8c,
        "M8": m8c,
        "N8": n8c,
        "O8": o8c,
        "P8": p8c,
    }
    formula_cached: Dict[str, float] = {**row8_cached, **dn_formula_cached}

    # --- 4) Summary sheet sync ---
    # Keep Summary formulas explicitly aligned with data-sheet dynamic cells.
    try:
        if "Summary" in wb.sheetnames:
            ws_summary = wb["Summary"]
            summary_links = {
                # Actuals section
                "A4": "='data-sheet'!A8",
                "B4": "='data-sheet'!D8",
                "C4": "='data-sheet'!E8",
                "D4": "='data-sheet'!F8",
                "E4": "='data-sheet'!J8",
                "G4": "='data-sheet'!K8",
                "H4": "='data-sheet'!L8",
                "I4": "='data-sheet'!M8",
                "J4": "='data-sheet'!N8",
                "K4": "='data-sheet'!O8",
                "L4": "='data-sheet'!P8",
                # Projection section
                "A7": "='data-sheet'!D12",
                "B7": "='data-sheet'!E12",
                "C7": "='data-sheet'!F12",
                "D7": "='data-sheet'!J12",
                "G7": "='data-sheet'!K12",
                "I7": "='data-sheet'!M12",
                "J7": "='data-sheet'!N12",
                "A8": "='data-sheet'!A14",
                "B8": "='data-sheet'!D14",
                "C8": "='data-sheet'!E14",
                "D8": "='data-sheet'!F14",
                "E8": "='data-sheet'!J14",
                "G8": "='data-sheet'!K14",
                "H8": "='data-sheet'!L14",
                "I8": "='data-sheet'!M14",
                "J8": "='data-sheet'!N14",
                "A9": "='data-sheet'!A15",
                "B9": "='data-sheet'!D15",
                "C9": "='data-sheet'!E15",
                "D9": "='data-sheet'!F15",
                "E9": "='data-sheet'!J15",
                "G9": "='data-sheet'!K15",
                "H9": "='data-sheet'!L15",
                "I9": "='data-sheet'!M15",
                "J9": "='data-sheet'!N15",
                "A10": "='data-sheet'!A16",
                "B10": "='data-sheet'!D16",
                "C10": "='data-sheet'!E16",
                "D10": "='data-sheet'!F16",
                "E10": "='data-sheet'!J16",
                "F10": "='data-sheet'!K16",
                "I10": "='data-sheet'!M16",
                "J10": "='data-sheet'!N16",
            }
            for addr, formula in summary_links.items():
                ws_summary[addr].value = formula
    except Exception:
        pass

    # Recalc on open so template formulas (M15, L4, J4, etc.) display correctly.
    # Previously this caused flicker because formulas had #REF! and data cells were text;
    # both are now fixed (valid formulas + numeric data cells).
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass

    return wb, formula_cached


@app.get("/api/route-report")
def api_route_report(
    route_id_site_id: str = Query(..., alias="route_id_site_id"),
    modality: Optional[str] = Query(None),
):
    """Return combined route report data as JSON for on-screen table."""
    rid = (route_id_site_id or "").strip()
    warnings: List[str] = []

    budget_rows = local_db.query_budget_by_site_id_all(rid)
    if not budget_rows:
        budget_rows = local_db.query_budget_by_route_id_insensitive(rid)
    if not budget_rows:
        warnings.append("No Budget data found for this route.")

    po_row = local_db.query_po_by_site_id(rid)
    if not po_row:
        po_row = local_db.query_po_by_route_id_insensitive(rid)
    if not po_row:
        warnings.append("No PO data found for this route.")

    dn_rows = local_db.get_dn_by_route_id_site_id(rid)
    if not dn_rows:
        dn_rows = local_db.get_dn_by_route_id_site_id_insensitive(rid)
    if not dn_rows:
        warnings.append("No DN data found for this route.")

    rows = _build_route_report_rows(route_id_site_id)

    wb, _formula_cached = _build_route_report_workbook(route_id_site_id, modality=modality)
    payload = _extract_route_report_summary_projection(wb)
    return {"rows": rows, "modality": modality, "warnings": warnings, **payload}


@app.get("/api/route-report/xlsx")
def api_route_report_xlsx(
    route_id_site_id: str = Query(..., alias="route_id_site_id"),
    modality: Optional[str] = Query(None),
):
    """Return downloadable Excel route report based on template."""
    import tempfile as _tempfile

    wb, formula_cached = _build_route_report_workbook(route_id_site_id, modality=modality)
    tmp = _tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.close()
    _patch_xlsx_formula_cached_values(tmp.name, "data-sheet", formula_cached)
    rid = (route_id_site_id or "").strip() or "route"
    filename = f"{rid}.xlsx"
    return FileResponse(
        tmp.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
    )


def _extract_route_report_summary_projection(wb: "Any") -> Dict[str, Any]:
    """
    Extract the "Summary" sheet values into two grids:
    - rows 2..5  => summary section
    - rows 6..10 => projection section

    The Summary sheet cells are mostly 1-level formulas that reference
    data-sheet cells (and occasionally internal Summary cell references).
    We also evaluate the referenced data-sheet formulas to provide
    Excel-matching numeric values on the UI.
    """
    import re as _re
    import openpyxl as _openpyxl
    from openpyxl.utils import get_column_letter as _col_letter

    ws_summary = wb["Summary"]
    ws_data = wb["data-sheet"]

    # ------------------------------
    # Helpers
    # ------------------------------
    def _col_to_idx(col: str) -> int:
        idx = 0
        for ch in col.upper():
            idx = idx * 26 + (ord(ch) - 64)
        return idx

    def _idx_to_col(idx: int) -> str:
        out = ""
        n = idx
        while n > 0:
            n, rem = divmod(n - 1, 26)
            out = chr(65 + rem) + out
        return out

    def _to_json_value(v: Any) -> Any:
        # Keep strings as-is so the UI shows labels correctly.
        if v is None:
            return ""
        # openpyxl uses datetime objects for dates.
        try:
            import datetime as _dt

            if isinstance(v, (_dt.date, _dt.datetime)):
                return v.date().isoformat()
        except Exception:
            pass
        # Avoid noisy floats in UI.
        if isinstance(v, float):
            # If it's effectively an integer, show as int.
            if abs(v - round(v)) < 1e-9:
                return int(round(v))
            return v
        return v

    # ------------------------------
    # Formula evaluation (data-sheet)
    # ------------------------------
    cache_data_num: Dict[str, float] = {}
    cache_data_raw: Dict[str, Any] = {}

    def _eval_cell_as_number(cell_addr: str, stack: "set[str]") -> float:
        """
        Evaluate a data-sheet cell to a number (used in arithmetic formulas).
        Non-numeric labels become 0.
        """
        cell_addr = cell_addr.upper().replace("$", "")
        if cell_addr in cache_data_num:
            return cache_data_num[cell_addr]
        if cell_addr in stack:
            # Cycle guard (shouldn't happen in this template).
            return 0.0
        stack.add(cell_addr)

        v = ws_data[cell_addr].value
        if v is None or v == "":
            out = 0.0
        elif isinstance(v, (int, float)):
            out = float(v)
        elif isinstance(v, str) and v.startswith("="):
            out = _eval_formula_to_number(v[1:], stack=stack)
        else:
            try:
                out = float(str(v).replace(",", ""))
            except Exception:
                out = 0.0

        cache_data_num[cell_addr] = out
        stack.remove(cell_addr)
        return out

    def _eval_range_sum(range_a: str, range_b: str, stack: "set[str]") -> float:
        # Both ends are like A1, B5.
        m1 = _re.fullmatch(r"([A-Z]{1,3})([0-9]+)", range_a.upper().replace("$", ""))
        m2 = _re.fullmatch(r"([A-Z]{1,3})([0-9]+)", range_b.upper().replace("$", ""))
        if not (m1 and m2):
            return 0.0
        c1, r1 = m1.group(1), int(m1.group(2))
        c2, r2 = m2.group(1), int(m2.group(2))
        c1i, c2i = _col_to_idx(c1), _col_to_idx(c2)
        rmin, rmax = min(r1, r2), max(r1, r2)
        cmin, cmax = min(c1i, c2i), max(c1i, c2i)
        s = 0.0
        for rr in range(rmin, rmax + 1):
            for cc in range(cmin, cmax + 1):
                addr = _idx_to_col(cc) + str(rr)
                s += _eval_cell_as_number(addr, stack)
        return s

    def _split_excel_args(s: str) -> List[str]:
        """
        Split function args by commas at top-level only.
        Example: SUM(A1, (B2+C2), D3) => ["A1","(B2+C2)","D3"]
        """
        args: List[str] = []
        depth = 0
        start = 0
        for i, ch in enumerate(s):
            if ch == "(":
                depth += 1
            elif ch == ")":
                depth = max(0, depth - 1)
            elif ch == "," and depth == 0:
                args.append(s[start:i].strip())
                start = i + 1
        tail = s[start:].strip()
        if tail:
            args.append(tail)
        return args

    def _eval_expression_to_number(expr: str, stack: "set[str]") -> float:
        expr = expr.strip()
        expr = expr.replace("$", "")
        if expr.startswith("(") and expr.endswith(")"):
            # keep parentheses for eval
            pass

        # Convert Excel '^' to python exponent if any.
        expr = expr.replace("^", "**")

        cell_tokens = set(_re.findall(r"\b[A-Z]{1,3}[0-9]{1,7}\b", expr.upper()))
        # Replace cell tokens with numeric values.
        for tok in cell_tokens:
            val = _eval_cell_as_number(tok, stack)
            # Use a stable repr for eval.
            expr = _re.sub(rf"\b{_re.escape(tok)}\b", repr(val), expr)

        try:
            return float(eval(expr, {"__builtins__": {}}, {}))
        except Exception:
            return 0.0

    def _eval_formula_to_number(formula_body: str, stack: "set[str]") -> float:
        body = formula_body.strip()
        body_u = body.upper()

        # SUM(...)
        if body_u.startswith("SUM(") and body.endswith(")"):
            inner = body[len("SUM(") : -1]
            # SUM may have a single arg like "O21+O30+..."
            args = _split_excel_args(inner)
            total = 0.0
            for arg in args:
                arg = arg.strip()
                m = _re.fullmatch(r"([A-Z]{1,3}[0-9]+):([A-Z]{1,3}[0-9]+)", arg.upper().replace("$", ""))
                if m:
                    total += _eval_range_sum(m.group(1), m.group(2), stack)
                else:
                    total += _eval_expression_to_number(arg, stack)
            return total

        # COUNTA(...) — non-empty cells (formula cells count as non-empty, like Excel).
        if body_u.startswith("COUNTA(") and body.endswith(")"):
            inner = body[len("COUNTA(") : -1]
            args = _split_excel_args(inner)
            count = 0.0

            def _counta_nonempty_raw(raw: Any) -> bool:
                if raw is None:
                    return False
                if isinstance(raw, str) and raw.startswith("="):
                    return True
                return str(raw).strip() != ""

            for arg in args:
                arg = arg.strip()
                m = _re.fullmatch(r"([A-Z]{1,3}[0-9]+):([A-Z]{1,3}[0-9]+)", arg.upper().replace("$", ""))
                if m:
                    m1 = _re.fullmatch(r"([A-Z]{1,3})([0-9]+)", m.group(1))
                    m2 = _re.fullmatch(r"([A-Z]{1,3})([0-9]+)", m.group(2))
                    if not (m1 and m2):
                        continue
                    c1, r1 = m1.group(1), int(m1.group(2))
                    c2, r2 = m2.group(1), int(m2.group(2))
                    c1i, c2i = _col_to_idx(c1), _col_to_idx(c2)
                    rmin, rmax = min(r1, r2), max(r1, r2)
                    cmin, cmax = min(c1i, c2i), max(c1i, c2i)
                    for rr in range(rmin, rmax + 1):
                        for cc in range(cmin, cmax + 1):
                            addr = _idx_to_col(cc) + str(rr)
                            if _counta_nonempty_raw(ws_data[addr].value):
                                count += 1.0
                elif _re.fullmatch(r"[A-Z]{1,3}[0-9]+", arg.upper().replace("$", "")):
                    addr = arg.upper().replace("$", "")
                    if _counta_nonempty_raw(ws_data[addr].value):
                        count += 1.0
            return count

        # IF(cond, true, false) — supports simple equality like D8=0 (template row 8 N8).
        if body_u.startswith("IF(") and body.endswith(")"):
            inner = body[len("IF(") : -1]
            args = _split_excel_args(inner)
            if len(args) >= 3:
                cond_s = args[0].strip()
                true_s = args[1].strip()
                false_s = args[2].strip()
                if "=" in cond_s:
                    left_e, right_e = cond_s.split("=", 1)
                    lv = _eval_expression_to_number(left_e.strip(), stack)
                    rv = _eval_expression_to_number(right_e.strip(), stack)
                    if abs(lv - rv) < 1e-12:
                        return _eval_expression_to_number(true_s, stack)
                    return _eval_expression_to_number(false_s, stack)

        # SUBTOTAL(3, ...)
        if body_u.startswith("SUBTOTAL(") and body.endswith(")"):
            inner = body[len("SUBTOTAL(") : -1]
            args = _split_excel_args(inner)
            if not args:
                return 0.0
            # args[0] is function_num. In this template, SUBTOTAL(3,...) means COUNTA.
            fn_raw = args[0].strip()
            try:
                fn_num = int(float(fn_raw))
            except Exception:
                fn_num = -1
            rest = args[1:]
            if fn_num == 3:
                count = 0.0
                for arg in rest:
                    arg = arg.strip()
                    m = _re.fullmatch(r"([A-Z]{1,3}[0-9]+):([A-Z]{1,3}[0-9]+)", arg.upper().replace("$", ""))
                    if m:
                        m1 = _re.fullmatch(r"([A-Z]{1,3})([0-9]+)", m.group(1))
                        m2 = _re.fullmatch(r"([A-Z]{1,3})([0-9]+)", m.group(2))
                        if not (m1 and m2):
                            continue
                        c1, r1 = m1.group(1), int(m1.group(2))
                        c2, r2 = m2.group(1), int(m2.group(2))
                        c1i, c2i = _col_to_idx(c1), _col_to_idx(c2)
                        rmin, rmax = min(r1, r2), max(r1, r2)
                        cmin, cmax = min(c1i, c2i), max(c1i, c2i)
                        for rr in range(rmin, rmax + 1):
                            for cc in range(cmin, cmax + 1):
                                addr = _idx_to_col(cc) + str(rr)
                                raw = ws_data[addr].value
                                if raw is not None and str(raw).strip() != "":
                                    count += 1.0
                    elif _re.fullmatch(r"[A-Z]{1,3}[0-9]+", arg.upper().replace("$", "")):
                        addr = arg.upper().replace("$", "")
                        raw = ws_data[addr].value
                        if raw is not None and str(raw).strip() != "":
                            count += 1.0
                    else:
                        # Fallback: treat non-zero expression as one non-empty value.
                        if _eval_expression_to_number(arg, stack) != 0:
                            count += 1.0
                return count

            # Fallback for other SUBTOTAL types: sum-like behavior.
            total = 0.0
            for arg in rest:
                arg = arg.strip()
                m = _re.fullmatch(r"([A-Z]{1,3}[0-9]+):([A-Z]{1,3}[0-9]+)", arg.upper().replace("$", ""))
                if m:
                    total += _eval_range_sum(m.group(1), m.group(2), stack)
                else:
                    total += _eval_expression_to_number(arg, stack)
            return total

        # Default: arithmetic expression
        return _eval_expression_to_number(body, stack)

    def _eval_data_raw(cell_addr: str) -> Any:
        """
        Evaluate a data-sheet cell into either a string/label (if not formula)
        or a numeric result (if formula).
        """
        cell_addr = cell_addr.upper().replace("$", "")
        if cell_addr in cache_data_raw:
            return cache_data_raw[cell_addr]
        v = ws_data[cell_addr].value
        if isinstance(v, str) and v.startswith("="):
            num = _eval_cell_as_number(cell_addr, set())
            # return numeric for UI.
            cache_data_raw[cell_addr] = num
        else:
            cache_data_raw[cell_addr] = v
        return cache_data_raw[cell_addr]

    # ------------------------------
    # Evaluate Summary cells
    # ------------------------------
    cache_summary_raw: Dict[str, Any] = {}

    def _eval_summary_cell(coord: str, stack: "set[str]") -> Any:
        coord = coord.upper()
        if coord in cache_summary_raw:
            return cache_summary_raw[coord]
        if coord in stack:
            return ""
        stack.add(coord)

        v = ws_summary[coord].value
        if isinstance(v, str) and v.startswith("='data-sheet'!"):
            addr = v.split("!", 1)[1].upper()
            out = _eval_data_raw(addr)
        elif isinstance(v, str) and v.startswith("="):
            ref = v[1:].strip().upper().replace("$", "")
            if _re.fullmatch(r"[A-Z]{1,3}[0-9]{1,7}", ref):
                out = _eval_summary_cell(ref, stack)
            else:
                # Unknown formula in Summary -> best effort.
                out = v
        else:
            out = v

        cache_summary_raw[coord] = out
        stack.remove(coord)
        return out

    def _make_grid(row_start: int, row_end: int) -> List[List[Any]]:
        grid: List[List[Any]] = []
        for r in range(row_start, row_end + 1):
            row: List[Any] = []
            for c in range(1, 13):  # A..L
                coord = _col_letter(c) + str(r)
                val = _eval_summary_cell(coord, set())
                row.append(_to_json_value(val))
            grid.append(row)
        return grid

    summary_grid = _make_grid(2, 5)
    projection_grid = _make_grid(6, 10)
    return {"summaryGrid": summary_grid, "projectionGrid": projection_grid}


# Field name from frontend validation table -> dn_master column
SEND_TO_MASTER_DN_FIELD_MAP = {
    "dn_number": "dn_number",
    "route_id_site_id": "route_id_site_id",
    "dn_length_mtr": "dn_length_mtr",
    "application_date": "application_date",
    "dn_received_date": "dn_received_date",
    "ground_rent": "ground_rent",
    "administrative_charge": "administrative_charge",
    "actual_total_non_refundable": "actual_total_non_refundable",
    "deposit": "deposit",
    "total_dn_amount": "total_dn_amount",
    "gst": "gst",
    "surface": "surface",
    "surface_wise_length": "surface_wise_length",
    "surface_wise_ri_amount": "surface_wise_ri_amount",
    "surface_wise_multiplication_factor": "surface_wise_multiplication_factor",
    "no_of_pits": "no_of_pits",
    "pit_ri_rate": "pit_ri_rate",
    "ot_length": "ot_length",
    "dn_ri_amount": "dn_ri_amount",
    "supervision_charges": "supervision_charges",
    "chamber_fee": "chamber_fee",
    "hdd_length": "hdd_length",
    "build_type": "build_type",
    "category_type": "category_type",
}


@app.post("/api/send-to-master-dn")
def send_to_master_dn(body: Dict[str, Any] = Body(...)):
    """Accept validation table data and upsert into dn_master."""
    data = body.get("data")  # list of { field, value }
    if not data:
        raise HTTPException(status_code=400, detail="Missing 'data' array")
    row = {}
    for item in data:
        field = (item.get("field") or "").strip()
        value = item.get("value")
        if field in SEND_TO_MASTER_DN_FIELD_MAP:
            col = SEND_TO_MASTER_DN_FIELD_MAP[field]
            if value is not None and str(value).strip() != "":
                v = str(value).strip()
                if col == "surface":
                    # Normalize parser-delimited values to comma-separated storage.
                    # Example: "A / B + C" -> "A, B, C"
                    import re as _re
                    v = _re.sub(r"\s*[\/+]\s*", ", ", v)
                    v = _re.sub(r"\s*,\s*", ", ", v).strip(" ,")
                if col in ("ground_rent", "gst", "deposit", "total_dn_amount", "actual_total_non_refundable",
                           "dn_length_mtr", "ot_length", "dn_ri_amount", "administrative_charge", "supervision_charges", "chamber_fee", "hdd_length", "pit_ri_rate"):
                    try:
                        row[col] = float(v)
                    except (ValueError, TypeError):
                        row[col] = v
                else:
                    row[col] = v
    if "dn_number" not in row:
        raise HTTPException(status_code=400, detail="dn_number is required")
    existing = local_db.get_dn_by_number(str(row["dn_number"]))
    if existing and existing.get("route_id_site_id") and row.get("route_id_site_id") and existing.get("route_id_site_id") != row.get("route_id_site_id"):
        raise HTTPException(status_code=409, detail="A Demand Note with this number already exists in the master database.")
    local_db.upsert_dn_master(row)
    after = None
    try:
        after = local_db.get_dn_by_number(str(row.get("dn_number")))
    except Exception:
        after = None
    return {"success": True, "message": "Saved to Master DN.", "after": after}


def _excel_to_dict_list(path: str, sheet_name=0):
    import pandas as pd
    df = pd.read_excel(path, sheet_name=sheet_name)
    df = df.where(df.notna(), None)
    return df.to_dict(orient="records")


def _normalize_col(name: str) -> str:
    return str(name).strip().lower().replace(" ", "_").replace("-", "_").replace("/", "_")


def _map_row_to_dn(row: Dict) -> Dict:
    """Map Excel row keys to dn_master columns."""
    allowed = {"dn_number", "route_id_site_id", "dn_length_mtr", "permission_receipt_date", "permit_no",
               "permit_start_date", "permit_end_date", "permitted_length_by_ward_mts", "surface", "no_of_pits",
               "ground_rent", "gst", "deposit", "total_dn_amount", "application_date", "dn_received_date",
               "actual_total_non_refundable", "po_number", "pit_ri_rate", "ot_length", "dn_ri_amount",
               "administrative_charge", "supervision_charges", "chamber_fee", "hdd_length", "build_type", "category_type"}
    out = {}
    for k, v in row.items():
        col = _normalize_col(k)
        if col in allowed and v is not None and str(v).strip() != "":
            out[col] = v
    return out


def _map_row_to_budget(row: Dict) -> Dict:
    allowed = {"route_id_site_id", "ce_length_mtr", "ri_cost_per_meter", "material_cost_per_meter", "build_cost_per_meter",
               "total_ri_amount", "material_cost", "execution_cost_including_hh", "total_cost_without_deposit",
               "route_type", "survey_id", "existing_new", "build_type", "category_type"}
    out = {}
    for k, v in row.items():
        col = _normalize_col(k)
        if col in allowed:
            if v is None or (isinstance(v, str) and v.strip() == ""):
                out[col] = None
            elif col in ("ce_length_mtr", "ri_cost_per_meter", "material_cost_per_meter", "build_cost_per_meter",
                         "total_ri_amount", "material_cost", "execution_cost_including_hh", "total_cost_without_deposit"):
                try:
                    out[col] = float(v)
                except (ValueError, TypeError):
                    out[col] = None
            else:
                out[col] = v
    return out


@app.post("/api/upload-dn-master")
async def upload_dn_master(file: UploadFile = File(...)):
    suffix = os.path.splitext(file.filename or "")[1] or ".xlsx"
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    tmp.write(await file.read())
    tmp.close()
    try:
        rows = _excel_to_dict_list(tmp.name)
        errors = []
        for r in rows:
            mapped = _map_row_to_dn(r)
            if mapped.get("dn_number"):
                try:
                    local_db.upsert_dn_master(mapped)
                except Exception as e:
                    errors.append(str(e))
        return {"success": True, "message": "Upload successful! New and updated rows processed.", "errors": errors[:10]}
    finally:
        try:
            os.remove(tmp.name)
        except Exception:
            pass


@app.post("/api/upload-budget-master")
async def upload_budget_master(file: UploadFile = File(...)):
    suffix = os.path.splitext(file.filename or "")[1] or ".xlsx"
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    tmp.write(await file.read())
    tmp.close()
    try:
        rows = _excel_to_dict_list(tmp.name)
        mapped = [_map_row_to_budget(r) for r in rows if _map_row_to_budget(r).get("route_id_site_id")]
        if mapped:
            local_db.budget_delete_not_in_site_ids([m.get("route_id_site_id") for m in mapped])
            local_db.budget_insert_many(mapped)
        return {"success": True, "message": "Upload successful!"}
    finally:
        try:
            os.remove(tmp.name)
        except Exception:
            pass


def _map_row_to_po_flat(row: Dict) -> Dict:
    """Map Excel row to po_master columns (normalize header names)."""
    out = {}
    for k, v in row.items():
        col = _normalize_col(k)
        if col == "route_id_site_id":
            out["route_id_site_id"] = v
        elif col == "route_type":
            out["route_type"] = v
        elif col in ("po_no_ip1", "po_no_ip_1"):
            out["po_no_ip1"] = v
        elif col in ("po_no_cobuild", "po_no_cobuild"):
            out["po_no_cobuild"] = v
        elif col in ("po_length_ip1", "po_length_ip_1"):
            out["po_length_ip1"] = v
        elif col in ("po_length_cobuild",):
            out["po_length_cobuild"] = v
        elif "route" in col and "lm" in col:
            out["route_routeLM_metroLM_LMCStandalone"] = v
    return out


@app.post("/api/upload-po-master")
async def upload_po_master(file: UploadFile = File(...)):
    suffix = os.path.splitext(file.filename or "")[1] or ".xlsx"
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    tmp.write(await file.read())
    tmp.close()
    try:
        rows = _excel_to_dict_list(tmp.name)
        engine = local_db.get_engine()
        from sqlalchemy import text
        with engine.connect() as conn:
            conn.execute(text("DELETE FROM po_master"))
            conn.commit()
        for r in rows:
            mapped = _map_row_to_po_flat(r)
            if not mapped.get("route_id_site_id"):
                continue
            local_db.upsert_po_master(mapped)
        return {"success": True, "message": "Upload successful!"}
    finally:
        try:
            os.remove(tmp.name)
        except Exception:
            pass


# -------------------------------------------------------------------
# VALIDATE PARSERS (PO, Application, DN)
# -------------------------------------------------------------------
@app.post("/api/parse-po")
async def parse_po(site_id: str = Form(...), po_number_type: Optional[str] = Form(None)):
    """Return PO data for the given site_id from the database (for Validate Parsers)."""
    row = local_db.query_po_by_site_id(site_id)
    if not row:
        return {"PO No": "", "PO Length (Mtr)": "", "Category": "", "SiteID": site_id, "UID": "", "Parent Route Name / HH": ""}
    route_type = (row.get("route_type") or "").replace(" ", "").lower()
    rrs = (row.get("route_routeLM_metroLM_LMCStandalone") or "").replace(" ", "").lower()
    use_cobuild = (po_number_type or "").lower() == "co-built" or rrs in ("metrolm", "lmc(standalone)", "routelm") or (rrs != "route" and route_type in ("metrolm", "lmc(standalone)", "routelm"))
    if use_cobuild:
        po_no = row.get("po_no_cobuild") or ""
        po_length = row.get("po_length_cobuild")
    else:
        po_no = row.get("po_no_ip1") or ""
        po_length = row.get("po_length_ip1")
    return {
        "PO No": str(po_no) if po_no else "",
        "PO Length (Mtr)": str(po_length) if po_length is not None else "",
        "Category": str(row.get("route_type") or ""),
        "SiteID": site_id,
        "UID": "",
        "Parent Route Name / HH": str(row.get("route_routeLM_metroLM_LMCStandalone") or ""),
    }


def _extract_po_essentials_from_text(text: str) -> Dict[str, Any]:
    src = text or ""
    po_number = ""
    route_id_site_id = ""
    po_value = ""
    entries: List[Dict[str, str]] = []

    # PO number appears like ".../10004960" in the sample.
    m_po = re.search(r"\bPO\s*No\.?\s*:?\s*([^\n\r]+)", src, flags=re.IGNORECASE)
    if m_po:
        line = m_po.group(1).strip()
        m_digits = re.search(r"(\d{6,12})(?!.*\d)", line)
        if m_digits:
            po_number = m_digits.group(1)
    if not po_number:
        m_fallback_po = re.search(r"/(\d{6,12})\b", src)
        if m_fallback_po:
            po_number = m_fallback_po.group(1)

    # Route/Site id in sample appears as MA4640.
    m_route = re.search(r"\b[A-Z]{2}\d{3,6}\b", src)
    if m_route:
        route_id_site_id = m_route.group(0)

    def _looks_like_route_site_id(line: str) -> bool:
        s = (line or "").strip()
        if not s:
            return False
        s_low = s.lower()
        if any(k in s_low for k in ("chapter heading", "hsn number", "sac number", "reference gbpa", "po no", "partner name")):
            return False
        if re.fullmatch(r"[A-Z]{2}\d{3,6}", s):
            return True
        # Numeric-only site ids are usually short (e.g., 2475), avoid long refs like 100016.
        if re.fullmatch(r"\d{3,5}", s):
            return True
        if "/" in s and re.search(r"\broute\b", s, flags=re.IGNORECASE):
            return True
        if re.search(r"\broute\b", s, flags=re.IGNORECASE):
            return True
        return False

    def _extract_route_for_item(window_text: str, fallback: str) -> str:
        lines = [ln.strip() for ln in (window_text or "").splitlines()]
        lines = [ln for ln in lines if ln]
        if not lines:
            return fallback

        chapter_idxs = [i for i, ln in enumerate(lines) if re.search(r"chapter\s*heading", ln, flags=re.IGNORECASE)]
        if chapter_idxs:
            idx = chapter_idxs[-1]
            # Primary rule: Route/Site id is the nearest useful line above "Chapter Heading".
            def _is_noise(ln: str) -> bool:
                t = (ln or "").strip().lower()
                if not t:
                    return True
                return any(k in t for k in (
                    "chapter heading",
                    "hsn number",
                    "sac number",
                    "this line",
                    "reference gbpa",
                    "po no",
                    "partner name",
                    "purchase order continuation sheet",
                ))

            for j in range(idx - 1, max(-1, idx - 8), -1):
                cand = lines[j].strip()
                if _is_noise(cand):
                    continue
                # Prefer explicit route-like candidates, otherwise still accept nearest text line.
                if _looks_like_route_site_id(cand):
                    return cand
                if len(cand) >= 3 and len(cand) <= 80:
                    return cand

            # Secondary rule: explicit route-like line in the nearby context.
            for j in range(idx - 1, max(-1, idx - 7), -1):
                cand = lines[j].strip()
                if _looks_like_route_site_id(cand):
                    return cand

        # Fallback 1: last slash-route style token in local window.
        route_like = re.findall(r"([A-Za-z0-9][A-Za-z0-9\s]*/\s*Route\s*/\s*\d+[^\n\r]*)", window_text, flags=re.IGNORECASE)
        if route_like:
            return route_like[-1].strip()

        # Fallback 2: last compact code like MA4640 in local window.
        compact = re.findall(r"\b[A-Z]{2}\d{3,6}\b", window_text)
        if compact:
            return compact[-1].strip()

        return fallback

    # Extract all table line items from the full PDF text:
    # Qty -> UOM -> Unit Price -> Line Total
    # Keep this format-driven (no value-specific assumptions).
    line_item_pat = re.compile(
        r"\b(?P<qty>\d{1,7})\s*\n\s*(?P<uom>[A-Za-z][A-Za-z ./-]{1,24})\s*\n\s*(?P<unit>\d[\d,]{2,})\s*\n\s*(?P<line>\d[\d,]{4,})\b",
        flags=re.IGNORECASE,
    )
    for i, m in enumerate(line_item_pat.finditer(src), start=1):
        qty = re.sub(r"[^\d]", "", m.group("qty"))
        uom = m.group("uom")
        unit_price = re.sub(r"[^\d]", "", m.group("unit"))
        line_total = re.sub(r"[^\d]", "", m.group("line"))
        route_window = src[max(0, m.start() - 2500): m.start()]
        row_route = _extract_route_for_item(route_window, route_id_site_id)
        entries.append(
            {
                "sr_no": str(i),
                "po_number": po_number,
                "route_id_site_id": row_route,
                "qty": qty,
                "uom": uom,
                "unit_price": unit_price,
                "po_value": line_total,
            }
        )

    if entries:
        # Keep top-level po_value as aggregate of all extracted line totals.
        try:
            po_value = str(sum(int(e.get("po_value") or "0") for e in entries))
        except Exception:
            po_value = entries[0].get("po_value") or ""
    else:
        # Fallback: single value near heading.
        m_seq = re.search(
            r"\b\d{1,6}\s*\n\s*(?:Meter|Metre)\s*\n\s*\d[\d,]{2,}\s*\n\s*(\d[\d,]{4,})\b",
            src,
            flags=re.IGNORECASE,
        )
        if m_seq:
            po_value = re.sub(r"[^\d]", "", m_seq.group(1))
        else:
            m_heading = re.search(r"Line\s*Total", src, flags=re.IGNORECASE)
            if m_heading:
                window = src[m_heading.end(): m_heading.end() + 1500]
                nums = [int(n.replace(",", "")) for n in re.findall(r"\b\d[\d,]{4,}\b", window)]
                nums = [n for n in nums if n <= 500000000]
                if nums:
                    po_value = str(max(nums))

    unique_routes = sorted({(e.get("route_id_site_id") or "").strip() for e in entries if (e.get("route_id_site_id") or "").strip()})
    if unique_routes:
        route_id_site_id = ", ".join(unique_routes)

    return {
        "po_number": po_number,
        "route_id_site_id": route_id_site_id,
        "po_value": po_value,
        "entries": entries,
        "entry_count": len(entries),
    }


@app.post("/api/parse-po-pdf")
async def parse_po_pdf(file: UploadFile = File(...)):
    """Extract PO Number, Route ID/Site ID, and PO Value from PO PDF."""
    suffix = os.path.splitext(file.filename or "")[1] or ".pdf"
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    tmp.write(await file.read())
    tmp.close()
    try:
        import fitz

        doc = fitz.open(tmp.name)
        text = "\n".join(page.get_text("text") for page in doc)
        doc.close()
        extracted = _extract_po_essentials_from_text(text)
        return extracted
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"PO PDF parsing failed: {e}")
    finally:
        try:
            os.remove(tmp.name)
        except Exception:
            pass


@app.post("/api/parse-application")
async def parse_application(dn_application_file: UploadFile = File(...), authority: str = Form(...)):
    """Parse DN Application PDF and return extracted fields."""
    suffix = os.path.splitext(dn_application_file.filename or "")[1] or ".pdf"
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    tmp.write(await dn_application_file.read())
    tmp.close()
    try:
        from parsers.universal_application_parser import universal_application_parser
        result = universal_application_parser(tmp.name)
        return result if isinstance(result, dict) else {}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        try:
            os.remove(tmp.name)
        except Exception:
            pass


def _run_dn_parser(pdf_path: str, authority: str) -> Dict[str, Any]:
    auth_lower = (authority or "").strip().upper()
    if auth_lower == "MCGM":
        from parsers.mcgm import non_refundable_request_parser
        return non_refundable_request_parser(pdf_path)
    if auth_lower == "MBMC":
        from parsers.mbmc import non_refundable_request_parser
        return non_refundable_request_parser(pdf_path)
    if auth_lower == "NMMC":
        from parsers.nmmc import non_refundable_request_parser
        return non_refundable_request_parser(pdf_path)
    if auth_lower == "KDMC":
        from parsers.kdmc import non_refundable_request_parser
        return non_refundable_request_parser(pdf_path)
    from parsers.mbmc import non_refundable_request_parser
    return non_refundable_request_parser(pdf_path)


@app.post("/api/parse-dn")
async def parse_dn(dn_file: UploadFile = File(...), authority: str = Form(...), site_id: Optional[str] = Form(None)):
    """Parse Demand Note PDF and return extracted fields."""
    suffix = os.path.splitext(dn_file.filename or "")[1] or ".pdf"
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    tmp.write(await dn_file.read())
    tmp.close()
    try:
        result = _run_dn_parser(tmp.name, authority)
        return result if isinstance(result, dict) else {}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        try:
            os.remove(tmp.name)
        except Exception:
            pass


# -------------------------------------------------------------------
# EXISTING APIs
# -------------------------------------------------------------------
@app.post("/api/nmmc-translate")
async def nmmc_translate(file: UploadFile = File(...)):
    suffix = os.path.splitext(file.filename or "")[1] or ".pdf"
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    tmp.write(await file.read())
    tmp.close()

    try:
        translated_text = nmmc.translate_pdf_to_english(tmp.name)
        return {"translated_text": translated_text}
    finally:
        try:
            os.remove(tmp.name)
        except Exception:
            pass


@app.post("/api/debug-mcgm-application")
async def debug_mcgm_application(file: UploadFile = File(...)):
    temp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    temp.write(await file.read())
    temp.close()

    try:
        doc = mcgm_application_parser.fitz.open(temp.name)
        text = "\n".join(page.get_text() for page in doc)
        doc.close()
        return {"text": text}
    finally:
        try:
            os.remove(temp.name)
        except Exception:
            pass
