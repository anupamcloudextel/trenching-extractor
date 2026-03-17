"""
Excel parsing and cleaning utilities for backend use.
"""
import shutil
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def clean_excel_file(input_path: str, output_path: str) -> str:
    """
    Cleans the Excel file according to the following logic:
    - For each row, if BOTH Length and Capex are blank, zero, or not > 0, set ID to blank and highlight red.
    - Otherwise, assign ID as MUMU[FY last 2 digits]RO[nnn], where nnn is a unique, zero-padded number within each FY group.
    Args:
        input_path (str): Path to the Excel file to clean.
        output_path (str): Path to write the cleaned Excel file.
    Returns:
        str: Status message.
    """
    wb = load_workbook(input_path)
    if 'ID Sheet' not in wb.sheetnames:
        wb.save(output_path)
        return "Sheet 'ID Sheet' not found. No changes made."
    ws = wb['ID Sheet']
    if ws is None:
        shutil.copyfile(input_path, output_path)
        return "No worksheet found. No changes made."

    # Find column indices
    headers = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}
    fy_col = headers.get("FY")
    length_col = headers.get("Length")
    capex_col = headers.get("Capex")
    id_col = headers.get("ID")
    wip_col = headers.get("WIP")
    if not (fy_col and length_col and capex_col and id_col):
        wb.save(output_path)
        return "Required columns not found. No changes made."

    # Red fill for highlighting
    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

    def is_blank_or_zero(val):
        if val is None:
            return True
        if isinstance(val, str):
            val = val.replace(',', '').strip()
        try:
            return float(val) <= 0
        except Exception:
            return True

    def is_invalid_fy(fy_val):
        if fy_val is None:
            return True
        if isinstance(fy_val, str):
            fy_val = fy_val.strip()
            if not fy_val:
                return True
            parts = fy_val.split('-')
            if len(parts) == 2 and parts[1].strip()[-2:].isdigit():
                return False  # Valid FY
            return True  # Invalid if can't extract last 2 digits
        return True  # Invalid if not a string or None

    # Step 1: Mark rows where BOTH Length and Capex are blank/zero/invalid, or FY is invalid, or WIP is not blank
    rows_to_skip = set()
    for row in range(2, ws.max_row + 1):
        length_val = ws.cell(row=row, column=length_col).value
        capex_val = ws.cell(row=row, column=capex_col).value
        fy_val = ws.cell(row=row, column=fy_col).value
        wip_val = ws.cell(row=row, column=wip_col).value if wip_col else None
        print(f"Row {row}: FY={fy_val!r}, Length={length_val!r}, Capex={capex_val!r}, WIP={wip_val!r}")
        if is_invalid_fy(fy_val):
            print(f"  -> Marked invalid: FY is blank, zero, or not in valid format.")
            cell = ws.cell(row=row, column=id_col)
            cell.value = ""
            cell.fill = red_fill
            rows_to_skip.add(row)
        elif is_blank_or_zero(length_val) and is_blank_or_zero(capex_val):
            print(f"  -> Marked invalid: Both Length and Capex are blank/zero/invalid.")
            cell = ws.cell(row=row, column=id_col)
            cell.value = ""
            cell.fill = red_fill
            rows_to_skip.add(row)
        elif wip_col and wip_val not in [None, "", 0]:
            print(f"  -> Marked invalid: WIP column is not blank.")
            cell = ws.cell(row=row, column=id_col)
            cell.value = ""
            cell.fill = red_fill
            rows_to_skip.add(row)

    # Step 2: Assign unique n within each FY group for remaining rows
    fy_to_rows = {}
    for row in range(2, ws.max_row + 1):
        if row in rows_to_skip:
            continue
        fy_val = ws.cell(row=row, column=fy_col).value
        if fy_val is not None:
            fy_to_rows.setdefault(fy_val, []).append(row)

    for fy, rows in fy_to_rows.items():
        for idx, r in enumerate(rows, 1):
            # Extract last 2 digits from FY (e.g., '2012-13' -> '13')
            fy_last2 = ""
            if fy and isinstance(fy, str):
                parts = fy.split('-')
                if len(parts) == 2 and parts[1]:
                    fy_last2 = parts[1].strip()[-2:]
            unique_str = f"{idx:03d}"
            id_val = f"MUMU{fy_last2}RO{unique_str}"
            print(f"Row {r}: Assigned ID={id_val}")
            ws.cell(row=r, column=id_col).value = id_val
            ws.cell(row=r, column=id_col).fill = PatternFill(fill_type=None)  # Remove fill if any

    wb.save(output_path)
    return "Excel file cleaned and duplicates processed." 